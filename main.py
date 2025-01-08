import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import os
import csv
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import fnmatch
import json
import re
from datetime import datetime

combined_list_pairs = [
    ("初回公表日", "初回公表日"),
    ("最終公表日", "最終公表日"),
    ("実施期間（終了日）", "*実施期間（終了日）"),
    ("研究の種別", "研究の種別"),
    ("治験の区分", "治験の区分"),
    (
        "責任研究者の組織名",
        "*Contact for Scientific Queries 所属機関（実施医療機関）",
        "*Contact for Scientific Queries 会社名・機関名",
    ),
    (
        "他施設の責任研究者の組織名",
        "*研究責任医師の連絡先*実施医療機関*",
        "*治験責任医師等の連絡先*実施医療機関*",
    ),
    (
        "多施設共同研究機関名",
        "*Affiliation",
    ),
    ("試験のフェーズ", "試験のフェーズ"),
    ("対象疾患名", "対象疾患名"),
    ("医薬品の一般名称", "医薬品等の一般名称"),
    ("販売名", "販売名", "*販売名*"),
    (
        "研究資金等の提供組織名称",
        "*依頼者以外の企業からの研究資金等の提供 研究資金等の提供組織名称",
    ),
    ("依頼者の名称", "*依頼者等に関する事項 依頼者等の名称"),
    ("他の臨床研究登録期間発行の研究番号", "*他の臨床研究登録機関発行の研究番号"),
    ("試験進捗状況", "進捗状況", "*進捗状況", "被験者募集状況"),
    ("JRCT_ID", "JRCT_ID"),
    ("研究名称", "*研究名称", "試験の名称"),
    ("組入れ開始日", "*実施期間（開始日）"),
    ("試験概要の目標症例数", "*目標症例数", "*実施予定被験者数*"),
    ("試験概要の試験のタイプ", "*試験の種類"),
    (
        "試験問い合わせ窓口のE-mail",
        "*Contact for Public Queries 電子メールアドレス",
        "*Contact for Public Queries 連絡先Eメールアドレス",
    ),
    ("試験問い合わせ窓口の担当者", "*Contact for Public Queries 担当者氏名"),
    ("url", "url"),
    ("平易な研究名称", "平易な研究名称"),
]


class ScrapingJRCT:
    def __init__(self, url, combined_list_pairs=None):
        self.url = url
        self.response = requests.get(url)
        self.soup = BeautifulSoup(self.response.content, "html.parser")
        self.combined_list_pairs = combined_list_pairs

    def get_first_table_data(self):
        # 一番初めのテーブル
        first_table = self.soup.find_all("table")[0]
        rows = first_table.find_all("tr")

        # 辞書を作成
        data = {}

        for row in rows:
            th = row.find("th")
            if th:
                label = th.get_text(strip=True)
                tds = row.find_all("td")
                values = [td.get_text(strip=True) for td in tds]

                # 辞書に追加
                if len(values) == 1:
                    data[label] = values[0]
                else:
                    data[label] = values
        return data

    def get_h2_h3_tags(self):
        nested_dict = {}
        current_h2 = None

        for tag in self.soup.find_all(["h2", "h3"]):
            if tag.name == "h2":
                current_h2 = tag.get_text(strip=True)
                nested_dict[current_h2] = []
            elif tag.name == "h3" and current_h2 is not None:
                nested_dict[current_h2].append(tag.get_text(strip=True))
        return nested_dict

    def find_h2_and_h3(self, h2_text, h3_text):
        h2_elements = self.soup.find_all("h2")
        for h2 in h2_elements:
            if h2_text == h2.get_text(strip=True):
                next_div = h2.find_next_sibling("div")
                h3_elements = next_div.find_all("h3")
                for h3 in h3_elements:
                    if h3.get_text(strip=True) == h3_text:
                        return h3

    def find_h2(self, h2_text):
        h2_elements = self.soup.find_all("h2")
        for h2 in h2_elements:
            if h2_text == h2.get_text(strip=True):
                return h2

    def find_next_table(self, h2or3_element):
        table = h2or3_element.find_next("table")
        return table

    def find_next_tables(self, h2or3_element):
        tables = []
        current_element = h2or3_element.find_next("table")
        while current_element and current_element.name == "table":
            tables.append(current_element)
            current_element = current_element.find_next_sibling()
            if current_element and current_element.name != "table":
                break
        return tables

    def table_to_dict(self, table):
        rows = table.find_all("tr")
        data = {}
        parent_label = ""
        for row in rows:
            th_s = row.find_all("th")
            if len(th_s) == 1:
                th = th_s[0]
            elif len(th_s) == 2:
                parent_label = th_s[0].get_text(strip=True)
                th = th_s[1]
            elif len(th_s) == 3:
                parent_label = (
                    th_s[0].get_text(strip=True) + " " + th_s[1].get_text(strip=True)
                )
                th = th_s[2]
            elif len(th_s) == 4:
                parent_label = (
                    th_s[0].get_text(strip=True)
                    + " "
                    + th_s[1].get_text(strip=True)
                    + " "
                    + th_s[2].get_text(strip=True)
                )
                th = th_s[3]

            label = th.get_text(strip=True)
            td_s = row.find_all("td")
            if len(td_s) == 2:
                # labelを変形
                # "/" で分割
                parts = label.split("/")
                if len(parts) == 2:
                    # 分割後の各部分の前後の空白を削除
                    japanese_label = parts[0].strip()
                    english_label = parts[1].strip()
                else:
                    japanese_label = parts[0].strip()
                    english_label = japanese_label
                if parent_label:
                    japanese_label = parent_label + " " + japanese_label
                    english_label = parent_label + " " + english_label
                data[japanese_label] = td_s[0].get_text(strip=True)
                data[english_label] = td_s[1].get_text(strip=True)
            elif len(td_s) == 1:
                if parent_label:
                    label = parent_label + " " + label
                data[label] = td_s[0].get_text(strip=True)

        return data

    def get_total_data(self):
        # tableがないなら
        if len(self.soup.find_all("table")) == 0:
            return dict()
        data = self.get_first_table_data()
        h2h3tags = self.get_h2_h3_tags()
        # 変更履歴は削除
        h2h3tags.pop("変更履歴", None)
        for h2, h3s in h2h3tags.items():
            if len(h3s) == 0:
                h2_element = self.find_h2(h2)
                tables = self.find_next_tables(h2_element)
                if tables:
                    for table in tables:
                        new_dic = self.table_to_dict(table)
                        for key, value in new_dic.items():
                            full_key = h2 + " " + key
                            # キーが既に存在するか確認
                            if full_key in data:
                                counter = 1
                                # キーが被らないように数字をつける
                                new_key = f"{full_key}_{counter}"
                                while new_key in data:
                                    counter += 1
                                    new_key = f"{full_key}_{counter}"
                                full_key = new_key
                            data[full_key] = value
            else:
                for h3 in h3s:
                    h3_element = self.find_h2_and_h3(h2, h3)
                    tables = self.find_next_tables(h3_element)
                    if tables:
                        for table in tables:
                            new_dic = self.table_to_dict(table)
                            for key, value in new_dic.items():
                                full_key = h2 + " " + key
                                # キーが既に存在するか確認
                                if full_key in data:
                                    counter = 1
                                    # キーが被らないように数字をつける
                                    new_key = f"{full_key}_{counter}"
                                    while new_key in data:
                                        counter += 1
                                        new_key = f"{full_key}_{counter}"
                                    full_key = new_key
                                data[full_key] = value
        return data


def add_column_to_csv(path, existing_columns):
    df = pd.read_csv(path)
    df = df.set_index("臨床研究実施計画番号")
    df = df.reindex(columns=existing_columns)
    df.to_csv(path, encoding="utf-8")


# 改行を削除する関数
def remove_newlines_from_keys(dictionary):
    cleaned_dict = {}
    for key, value in dictionary.items():
        new_key = key.replace("\n", " ").replace("\r", " ")
        cleaned_dict[new_key] = value
    return cleaned_dict


def remove_illegal_characters(value):
    if isinstance(value, str):
        # Excelで許可されない文字を削除
        return re.sub(r"[\x00-\x09\x0B-\x1F\x7F-\x9F]", "", value)
    return value


# 和暦の '令和X年Y月Z日' を西暦の datetime に変換する関数
def wareki_to_seireki(date_str):
    if "令和" in date_str:
        year, month, day = (
            date_str.replace("令和", "")
            .replace("年", "-")
            .replace("月", "-")
            .replace("日", "")
            .split("-")
        )
        if year == "元":
            year = 1
        year = int(year) + 2018  # 令和元年は2019年
    elif "平成" in date_str:
        year, month, day = (
            date_str.replace("平成", "")
            .replace("年", "-")
            .replace("月", "-")
            .replace("日", "")
            .split("-")
        )
        if year == "元":
            year = 1
        year = int(year) + 1988  # 平成元年は1989年
    elif "昭和" in date_str:
        year, month, day = (
            date_str.replace("昭和", "")
            .replace("年", "-")
            .replace("月", "-")
            .replace("日", "")
            .split("-")
        )
        if year == "元":
            year = 1
        year = int(year) + 1925  # 昭和元年は1926年
    else:
        raise ValueError(f"Unsupported era in date string: {date_str}")

    # 日付をdatetimeオブジェクトに変換して返す
    return datetime(int(year), int(month), int(day))


def update(index_list, excel_file_path, json_file_path):
    # JSONファイルを読み込む
    if os.path.exists(json_file_path):
        with open(json_file_path, "r") as file:
            json_data = json.load(file)
    else:
        json_data = dict()

    # excelファイルがないなら作成しヘッダーを書き込む
    if not os.path.exists(excel_file_path):
        # Excelファイルを新規作成
        workbook = Workbook()
        # シートを取得
        sheet = workbook.active
        sheet.title = "Sheet1"
        # ヘッダーを書き込む
        excel_column_names = [column[0] for column in combined_list_pairs]
        sheet.append(excel_column_names)
        # ファイルを保存
        workbook.save(excel_file_path)
    else:
        # Excelファイルを開く
        workbook = load_workbook(excel_file_path)

    # データフレームを読み込む
    existing_df = pd.read_excel(excel_file_path)
    existing_df = existing_df.set_index("JRCT_ID")
    existing_ids = set(existing_df.index)

    # 追記するシート名を指定（既存のシート名）
    sheet = workbook["Sheet1"]

    study_status_index = get_index_by_first_element("試験進捗状況", combined_list_pairs)

    for index in tqdm(index_list):
        url = "https://jrct.niph.go.jp/latest-detail/" + index
        # すでに取得済みのIDならスキップする
        if index in existing_ids:
            continue

        if index in json_data and len(json_data[index]) > 0:
            data = json_data[index]
        else:
            for _ in range(3):
                time.sleep(1)  # スクレイピングの間隔を開ける
                scraper = ScrapingJRCT(url)
                data = scraper.get_total_data()
                if len(data) == 0:
                    time.sleep(10)
                    continue
                # 辞書のキーから改行を削除
                data = remove_newlines_from_keys(data)
                json_data[index] = data
                break

            if len(data) == 0:
                print(
                    f"Failed to get data for {index}. 自動的に再実行されます．try again after 100 seconds."
                )
                time.sleep(100)
                continue

        new_row = []

        for column_names in combined_list_pairs:
            if column_names[0] == "JRCT_ID":
                new_row.append(index)
                continue
            if column_names[0] == "url":
                new_row.append(url)
                continue

            # ここでは複数とる
            if column_names[0] == "他施設の責任研究者の組織名":
                matching_columns = []
                for column_name in column_names[1:]:
                    add_matching_columns = [
                        match_column_name
                        for match_column_name in data.keys()
                        if fnmatch.fnmatch(match_column_name, column_name)
                    ]
                    matching_columns.extend(add_matching_columns)

                if len(matching_columns) == 0:
                    new_row.append("")
                else:
                    institute_names = ""
                    for key in matching_columns:
                        institute_names += data[key] + "\n"
                    new_row.append(institute_names)
                continue

            for column_name in column_names[1:]:
                matching_columns = [
                    match_column_name
                    for match_column_name in data.keys()
                    if fnmatch.fnmatch(match_column_name, column_name)
                ]

                if not len(matching_columns) == 0:
                    break

            if len(matching_columns) == 0:
                new_row.append("")
            else:
                key = matching_columns[0]
                new_row.append(data[key])

        new_row = [remove_illegal_characters(value) for value in new_row]

        existing_ids.add(index)
        # 終了の時は追記せずに飛ばす
        if (
            new_row[study_status_index] == "研究終了"
            or new_row[study_status_index] == "募集終了"
        ):
            continue

        # 追記
        sheet.append(new_row)

        if sheet.max_row % 10 == 0:
            workbook.save(excel_file_path)

        if len(json_data) % 10 == 0:
            with open(json_file_path, "w") as file:
                json.dump(json_data, file, indent=4, ensure_ascii=False)

    # 最後に残ったデータを追記
    workbook.save(excel_file_path)
    with open(json_file_path, "w") as file:
        json.dump(json_data, file, indent=4, ensure_ascii=False)

    # 最後にexcelファイルを並び替える
    # Excelファイルを読み込み
    df = pd.read_excel(excel_file_path, sheet_name="Sheet1")

    # 'Date' 列を和暦から西暦の datetime に変換
    df["date"] = df["初回公表日"].apply(wareki_to_seireki)

    # 日付の列で並び替え
    df_sorted = df.sort_values(by="date", ascending=False)

    # dateの列を削除
    df_sorted = df_sorted.drop(columns=["date"])

    # 並び替えたデータを新しいファイルに保存
    df_sorted.to_excel(excel_file_path, index=False)


def get_row_data(data):
    new_row = []
    for column_names in combined_list_pairs:
        if column_names[0] == "JRCT_ID":
            new_row.append("jrct_id")
            continue

        for column_name in column_names[1:]:
            matching_columns = [
                match_column_name
                for match_column_name in data.keys()
                if fnmatch.fnmatch(match_column_name, column_name)
            ]

            if not len(matching_columns) == 0:
                break

        if len(matching_columns) == 0:
            new_row.append("")
        else:
            key = matching_columns[0]
            new_row.append(data[key])
    return new_row


def get_index_by_first_element(element_name, combined_list_pairs):
    for index, pair in enumerate(combined_list_pairs):
        if pair[0] == element_name:
            return index
    return -1  # 該当なしの場合


def main():
    # indexのcsvを読み込む
    index_csv_file_path = "jrct_index_data.csv"
    index_df = pd.read_csv(
        index_csv_file_path, index_col="臨床研究実施計画番号", encoding="utf-8"
    )
    index_list = index_df.index.tolist()
    #index_list = ['jRCT2061240025']
    update(index_list, "jrct_data.xlsx", "jrct_data.json")


if __name__ == "__main__":
    main()
